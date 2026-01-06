# import streamlit as st
# import pandas as pd
# from io import StringIO

# st.set_page_config(page_title="CSV/Excel 表示アプリ（Streamlit版）", layout="wide")
# st.title("CSV/Excel 表示アプリ（Streamlit版）")

# uploaded_file = st.file_uploader("CSVまたはExcelファイルを選択", type=["csv", "xlsx"])

# if uploaded_file:
#     header_keywords = ['公報番号']
#     encodings = ['utf-8-sig', 'utf-8', 'cp932']
#     df = None

#     try:
#         if uploaded_file.name.lower().endswith(".csv"):
#             raw_bytes = uploaded_file.read()

#             for enc in encodings:
#                 try:
#                     text = raw_bytes.decode(enc)
#                 except Exception:
#                     continue

#                 lines = text.splitlines()
#                 header_index = None
#                 for i, line in enumerate(lines[:10]):  # 先頭10行まで探索
#                     if any(kw in line for kw in header_keywords):
#                         header_index = i
#                         break

#                 try:
#                     if header_index is not None:
#                         csv_text = "\n".join(lines[header_index:])
#                         df = pd.read_csv(StringIO(csv_text), sep=",", engine="python", dtype=str, on_bad_lines="skip", index_col=False)
#                     else:
#                         df = pd.read_csv(StringIO(text), sep=",", engine="python", dtype=str, on_bad_lines="skip", index_col=False)
#                     st.success(f"エンコーディング: {enc}, ヘッダー行: {header_index}")
#                     break
#                 except Exception as e:
#                     st.warning(f"{enc} での読み込み失敗: {e}")
#                     df = None
#                     continue

#             if df is None:
#                 uploaded_file.seek(0)  # ★ ここが重要（先頭に戻す）
#                 try:
#                     df = pd.read_csv(uploaded_file, dtype=str, on_bad_lines="skip", encoding=enc, engine="python")
#                 except Exception as e:
#                     st.error(f"最終読み込み失敗: {e}")
#                     df = pd.DataFrame()
#         else:
#             df = pd.read_excel(uploaded_file, dtype=str)

#         if df is not None:
#             df = df.fillna("")
#             # st.dataframe(df, width='content')
#             st.dataframe(df, use_container_width=True)

#     except Exception as ex:
#         st.error(f"読み込みエラー: {ex}")


import streamlit as st
import pandas as pd
from io import StringIO
from openpyxl import load_workbook

st.set_page_config(page_title="CSV/Excel 表示アプリ（両対応）", layout="wide")
st.title("CSV/Excel 表示アプリ（CSV / .xlsx）")

# 設定
HEADER_KEYWORDS = ("公報番号",)  # ヘッダー検出用キーワード
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp932")  # 試す順番
CSV_SEARCH_ROWS = 50  # CSVの先頭何行までヘッダ探索するか
XLSX_SEARCH_ROWS = 50  # Excelの先頭何行までヘッダ探索するか

uploaded_file = st.file_uploader("CSV または Excel (.xlsx) を選択", type=["csv", "xlsx"])

def read_csv_auto(file) -> tuple[pd.DataFrame, dict]:
    """
    CSVの自動読み込み:
      - エンコーディングを候補から順にデコード
      - 先頭N行で「公報番号」を含む行をヘッダ行として検出
      - 検出に応じてpd.read_csvに渡すテキストを調整
    戻り値: (DataFrame, メタ情報dict)
    """
    raw = file.read()  # bytes
    meta = {"encoding": None, "header_index": None, "source": "csv"}

    for enc in CSV_ENCODINGS:
        try:
            text = raw.decode(enc)
        except Exception:
            continue

        lines = text.splitlines()
        header_index = None
        # 先頭N行の中からヘッダー行を探す
        for i, line in enumerate(lines[:CSV_SEARCH_ROWS]):
            if any(kw in line for kw in HEADER_KEYWORDS):
                header_index = i
                break

        try:
            if header_index is not None:
                csv_text = "\n".join(lines[header_index:])
            else:
                csv_text = text

            df = pd.read_csv(
                StringIO(csv_text),
                sep=",",
                engine="python",
                dtype=str,
                on_bad_lines="skip",
                index_col=False,
            )
            meta["encoding"] = enc
            meta["header_index"] = header_index
            return df, meta
        except Exception:
            # 次のエンコーディングへ
            pass

    # すべて失敗した場合の最後のトライ（cp932固定など）
    try:
        text = raw.decode("cp932")
        df = pd.read_csv(
            StringIO(text),
            sep=",",
            engine="python",
            dtype=str,
            on_bad_lines="skip",
            index_col=False,
        )
        meta["encoding"] = "cp932"
        meta["header_index"] = None
        return df, meta
    except Exception as e:
        raise RuntimeError(f"CSV読み込み失敗: {e}")

def list_xlsx_sheets(file) -> list[str]:
    """openpyxlでシート名一覧を取得"""
    file.seek(0)
    wb = load_workbook(file, read_only=True, data_only=True)
    return wb.sheetnames

def find_header_row_xlsx(file_like, sheet_name=None) -> int | None:
    """
    openpyxlで先頭からXLSX_SEARCH_ROWS行を走査し、
    HEADER_KEYWORDSのいずれかを含むセルがある行番号(1始まり)を返す。なければNone。
    """
    file_like.seek(0)
    wb = load_workbook(file_like, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=XLSX_SEARCH_ROWS), start=1):
        values = [str(c.value) if c.value is not None else "" for c in row]
        joined = "\t".join(values)
        if any(kw in joined for kw in HEADER_KEYWORDS):
            return i  # 1-based
    return None

def read_xlsx_openpyxl(file, sheet_name=None) -> tuple[pd.DataFrame, dict]:
    """
    Excel(.xlsx)読み込み:
      - シート名指定可（未指定なら先頭）
      - 「公報番号」ヘッダ行検出（見つかればその行をヘッダに、見つからなければデフォルトheader=0）
    戻り値: (DataFrame, メタ情報dict)
    """
    header_row_1based = find_header_row_xlsx(file, sheet_name=sheet_name)
    # pandasのheaderは0始まり。見つからない場合は0（先頭行をヘッダ）にするのが無難
    header_arg = header_row_1based - 1 if header_row_1based is not None else 0

    file.seek(0)
    df = pd.read_excel(
        file,
        dtype=str,
        engine="openpyxl",
        sheet_name=sheet_name if sheet_name else 0,
        header=header_arg,
    )

    meta = {"sheet": sheet_name if sheet_name else "(先頭)", "header_row": header_row_1based, "source": "xlsx"}
    return df, meta

if uploaded_file:
    ext = uploaded_file.name.lower().rsplit(".", 1)[-1]
    try:
        if ext == "csv":
            df, meta = read_csv_auto(uploaded_file)
            df = df.fillna("")
            st.success(f"[CSV] エンコーディング: {meta['encoding']} / ヘッダー行: {meta['header_index']}")
            st.dataframe(df, use_container_width=True)

        elif ext == "xlsx":
            # まずシート名一覧を取得して選択させる
            sheet_names = list_xlsx_sheets(uploaded_file)
            sel = st.selectbox("読み込むシートを選択", options=sheet_names, index=0)
            df, meta = read_xlsx_openpyxl(uploaded_file, sheet_name=sel)
            df = df.fillna("")
            st.success(f"[XLSX] シート: {meta['sheet']} / ヘッダー行: {meta['header_row'] if meta['header_row'] is not None else '先頭行'}")
            st.dataframe(df, use_container_width=True)

        else:
            st.error("未対応の拡張子です。CSVまたは.xlsxを選択してください。")

    except Exception as ex:
        st.error(f"読み込みエラー: {ex}")
