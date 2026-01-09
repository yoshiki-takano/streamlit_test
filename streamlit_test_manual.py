# streamlit_test_manual.py
import streamlit as st
import pandas as pd
from io import StringIO
from openpyxl import load_workbook

st.set_page_config(page_title="CSV/Excel 表示アプリ（ヘッダー手動調整対応）", layout="wide",
                   initial_sidebar_state="collapsed")
st.title("CSV/Excel 表示アプリ（CSV / .xlsx）")

# ===== 設定（デフォルト） =====
DEFAULT_HEADER_KEYWORDS = ("公報番号",)
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp932")
CSV_SEARCH_ROWS = 50
XLSX_SEARCH_ROWS = 50

uploaded_file = st.file_uploader("CSV または Excel (.xlsx) を選択", type=["csv", "xlsx"]) 

# ===== ユーティリティ =====
def detect_csv(raw: bytes, encodings=CSV_ENCODINGS, search_rows=CSV_SEARCH_ROWS, header_keywords=DEFAULT_HEADER_KEYWORDS):
    """CSVの自動検出（エンコーディングとヘッダー行）"""
    meta = {"encoding": None, "header_index": None, "source": "csv"}
    for enc in encodings:
        try:
            text = raw.decode(enc)
        except Exception:
            continue
        lines = text.splitlines()
        header_index = None
        for i, line in enumerate(lines[:search_rows]):
            if any(kw in line for kw in header_keywords):
                header_index = i
                break
        try:
            csv_text = "\n".join(lines[header_index:]) if header_index is not None else text
            df = pd.read_csv(StringIO(csv_text), sep=",", engine="python", dtype=str, on_bad_lines="skip", index_col=False)
            meta["encoding"] = enc
            meta["header_index"] = header_index
            return df, meta
        except Exception:
            pass
    # すべて失敗したらcp932で最後のトライ
    try:
        text = raw.decode("cp932")
        df = pd.read_csv(StringIO(text), sep=",", engine="python", dtype=str, on_bad_lines="skip", index_col=False)
        meta["encoding"] = "cp932"
        meta["header_index"] = None
        return df, meta
    except Exception as e:
        raise RuntimeError(f"CSV読み込み失敗: {e}")

def read_csv_with_options(raw: bytes, encoding: str, header_row_1based: int | None):
    """指定のエンコーディング/ヘッダー行でCSVを読む。header_row_1basedがNoneなら先頭行から。"""
    text = raw.decode(encoding)
    lines = text.splitlines()
    if header_row_1based is not None:
        header_index = max(0, header_row_1based - 1)
        csv_text = "\n".join(lines[header_index:])
    else:
        csv_text = text
    df = pd.read_csv(StringIO(csv_text), sep=",", engine="python", dtype=str, on_bad_lines="skip", index_col=False)
    return df

def list_xlsx_sheets(file_like) -> list[str]:
    file_like.seek(0)
    wb = load_workbook(file_like, read_only=True, data_only=True)
    return wb.sheetnames

def find_header_row_xlsx(file_like, sheet_name=None, search_rows=XLSX_SEARCH_ROWS, header_keywords=DEFAULT_HEADER_KEYWORDS):
    """openpyxlで検索して、ヘッダー候補の行番号(1始まり)を返す。無ければNone。"""
    file_like.seek(0)
    wb = load_workbook(file_like, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=search_rows), start=1):
        values = [str(c.value) if c.value is not None else "" for c in row]
        joined = "\t".join(values)
        if any(kw in joined for kw in header_keywords):
            return i
    return None

def read_xlsx(file_like, sheet_name=None, header_row_1based: int | None = None):
    """ヘッダー行を指定してExcelを読む（未指定ならpandasのデフォルト=先頭行）。"""
    header_arg = header_row_1based - 1 if header_row_1based is not None else 0
    file_like.seek(0)
    df = pd.read_excel(
        file_like,
        dtype=str,
        engine="openpyxl",
        sheet_name=sheet_name if sheet_name else 0,
        header=header_arg,
    )
    return df

# ===== 画面構成 =====
with st.sidebar:
    st.markdown("### ヘッダー検出の調整")
    kw_text = st.text_input("検出キーワード（カンマ区切り）", value=",".join(DEFAULT_HEADER_KEYWORDS))
    header_keywords = tuple(x.strip() for x in kw_text.split(",") if x.strip()) or DEFAULT_HEADER_KEYWORDS
    st.caption("例: 公報番号, 出願番号 など")

if uploaded_file:
    ext = uploaded_file.name.lower().rsplit(".", 1)[-1]
    try:
        if ext == "csv":
            # まずは自動検出
            uploaded_file.seek(0)
            raw = uploaded_file.read()
            auto_df, meta = detect_csv(raw, header_keywords=header_keywords)
            st.success(f"[CSV] 自動検出: エンコーディング={meta['encoding']} / ヘッダー行={meta['header_index']} (0始まり)")

            # 手動調整UI
            with st.expander("ヘッダー・エンコーディングの手動調整（CSV）", expanded=False):
                enc_options = list(dict.fromkeys(list(CSV_ENCODINGS) + ["cp932", meta["encoding"] or "utf-8"]))
                default_enc_idx = enc_options.index(meta["encoding"]) if meta["encoding"] in enc_options else 0
                sel_enc = st.selectbox("エンコーディング", options=enc_options, index=default_enc_idx)

                use_manual_header = st.checkbox("ヘッダー行を手動指定する")
                default_header_1based = (meta["header_index"] + 1) if meta["header_index"] is not None else 1
                manual_header_1based = st.number_input("ヘッダー行（1始まり）", min_value=1, max_value=1000, value=default_header_1based)

                st.caption("先頭から何行目が列名（ヘッダー）かを指定します。")

            # 読み込み（手動指定があればそれを優先）
            header_arg = manual_header_1based if use_manual_header else None
            df = read_csv_with_options(raw, encoding=sel_enc, header_row_1based=header_arg)
            df = df.fillna("")

            # 先頭行プレビュー（選択の参考）
            with st.expander("CSV先頭プレビュー（行番号付き）"):
                text_preview = raw.decode(sel_enc, errors="replace")
                lines = text_preview.splitlines()[:200]
                numbered = [f"{i+1:>4}: " + line for i, line in enumerate(lines)]
                st.code("\n".join(numbered), language="text")

            st.dataframe(df, use_container_width=True)

        elif ext == "xlsx":
            # シート選択
            sheet_names = list_xlsx_sheets(uploaded_file)
            sel_sheet = st.selectbox("読み込むシートを選択", options=sheet_names, index=0)

            # 自動検出
            header_row_1based_auto = find_header_row_xlsx(uploaded_file, sheet_name=sel_sheet, header_keywords=header_keywords)
            st.success(f"[XLSX] 自動検出: シート={sel_sheet} / ヘッダー行={(header_row_1based_auto if header_row_1based_auto is not None else '先頭行')} ")

            # 手動調整UI
            with st.expander("ヘッダーの手動調整（Excel）", expanded=False):
                use_manual_header_x = st.checkbox("ヘッダー行を手動指定する (Excel)")
                default_header_x_1based = header_row_1based_auto if header_row_1based_auto is not None else 1
                manual_header_x_1based = st.number_input("ヘッダー行（1始まり）(Excel)", min_value=1, max_value=1000, value=default_header_x_1based)
                st.caption("指定した行が列名になります。")

            header_arg_x = manual_header_x_1based if use_manual_header_x else (header_row_1based_auto if header_row_1based_auto is not None else 1)
            df = read_xlsx(uploaded_file, sheet_name=sel_sheet, header_row_1based=header_arg_x)
            df = df.fillna("")
            st.dataframe(df, use_container_width=True)
        else:
            st.error("未対応の拡張子です。CSVまたは.xlsxを選択してください。")
    except Exception as ex:
        st.error(f"読み込みエラー: {ex}")
else:
    st.info("ファイルをアップロードしてください。")
